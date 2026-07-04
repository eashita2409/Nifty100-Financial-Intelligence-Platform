import sqlite3
import pandas as pd
import numpy as np
import os
from openpyxl.styles import PatternFill, Font
from openpyxl.formatting.rule import CellIsRule

# 1. Load Data
db_path = 'data/db/nifty100.db'
conn = sqlite3.connect(db_path)
peer_groups = pd.read_sql_query('SELECT company_id, peer_group_name, is_benchmark FROM peer_groups', conn)
conn.close()

df = pd.read_excel('output/screener_output.xlsx')

# 2. Select 20 KPIs
kpi_columns = [
    'composite_quality_score', 'revenue_growth', 'pat_growth', 'eps_growth', 
    'roe', 'roce', 'debt_equity', 'current_ratio', 'quick_ratio', 
    'interest_coverage', 'operating_margin', 'net_margin', 'cash_flow_growth', 
    'market_cap_growth', 'peg', 'pe', 'pb', 'dividend_yield', 
    'enterprise_value', 'fcf'
]

# Ensure we have company_name and company_id
df = df[['company_id', 'company_name'] + kpi_columns]

# Merge with peer groups
# We only care about companies that are in a peer group
merged = pd.merge(peer_groups, df, on='company_id', how='inner')

# Lower is better KPIs
lower_is_better = ['debt_equity', 'peg', 'pe', 'pb']

# Calculate universe percentiles for all KPIs
percentile_df = df.copy()
for col in kpi_columns:
    ascending = False if col in lower_is_better else True
    # Calculate rank as percentile 0-100
    percentile_col = f"{col}_pct"
    percentile_df[percentile_col] = df[col].rank(pct=True, ascending=ascending) * 100

# Merge percentiles back
merged = pd.merge(merged, percentile_df[['company_id'] + [f"{c}_pct" for c in kpi_columns]], on='company_id', how='left')

# Prepare Excel Writer
output_path = 'output/peer_comparison.xlsx'
writer = pd.ExcelWriter(output_path, engine='openpyxl')

groups = merged['peer_group_name'].unique()

for group in groups:
    group_df = merged[merged['peer_group_name'] == group].copy()
    
    # We want columns in order: company_name, is_benchmark, [KPI1, KPI1_pct, KPI2, KPI2_pct...]
    cols_to_display = ['company_name']
    for kpi in kpi_columns:
        cols_to_display.extend([kpi, f"{kpi}_pct"])
        
    display_df = group_df[cols_to_display].copy()
    
    # Add median row
    median_vals = display_df.drop(columns=['company_name']).median()
    median_row = pd.DataFrame([['Median'] + median_vals.tolist()], columns=cols_to_display)
    display_df = pd.concat([display_df, median_row], ignore_index=True)
    
    # Write to sheet
    sheet_name = str(group)[:31] # Excel limits sheet name to 31 chars
    display_df.to_excel(writer, sheet_name=sheet_name, index=False)
    
    worksheet = writer.sheets[sheet_name]
    
    # Apply formatting
    num_rows = len(display_df)
    
    # Formats
    gold_fill = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')
    median_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
    bold_font = Font(bold=True)
    
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    green_font = Font(color='006100')
    
    yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    yellow_font = Font(color='9C5700')
    
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    red_font = Font(color='9C0006')

    # Benchmark highlight
    benchmark_indices = group_df.reset_index().index[group_df['is_benchmark'] == 1].tolist()
    
    for row_num in range(num_rows):
        if row_num in benchmark_indices:
            # openpyxl is 1-indexed, header is row 1, data starts at row 2
            for cell in worksheet[row_num + 2]:
                cell.fill = gold_fill
                cell.font = bold_font
        elif row_num == num_rows - 1:
            # Median row
            for cell in worksheet[row_num + 2]:
                cell.fill = median_fill
                cell.font = bold_font
            
    # Conditional formatting for percentile columns
    from openpyxl.utils import get_column_letter
    
    for i, col in enumerate(cols_to_display):
        col_letter = get_column_letter(i + 1)
        
        # Set column width
        if col == 'company_name':
            worksheet.column_dimensions[col_letter].width = 25
        else:
            worksheet.column_dimensions[col_letter].width = 15
            
        if col.endswith('_pct'):
            # Apply conditional formatting to this column
            # Excluding header and median row
            range_str = f'{col_letter}2:{col_letter}{num_rows}' # num_rows includes median, so num_rows (index) is the last company, wait, num_rows is length of display_df. display_df has median. So num_rows+1 is the median row. We want to exclude median row, so up to num_rows. 
            
            # Green >= 75
            worksheet.conditional_formatting.add(range_str, CellIsRule(operator='greaterThanOrEqual', formula=['75'], stopIfTrue=True, fill=green_fill, font=green_font))
            
            # Yellow 25-75 (between 25 and 75)
            worksheet.conditional_formatting.add(range_str, CellIsRule(operator='between', formula=['25.0001', '74.9999'], stopIfTrue=True, fill=yellow_fill, font=yellow_font))
            
            # Red <= 25
            worksheet.conditional_formatting.add(range_str, CellIsRule(operator='lessThanOrEqual', formula=['25'], stopIfTrue=True, fill=red_fill, font=red_font))
            
            # Format numbers to 1 decimal
            for row in range(2, num_rows + 2):
                worksheet[f"{col_letter}{row}"].number_format = '0.0'
        else:
            if col != 'company_name':
                for row in range(2, num_rows + 2):
                    worksheet[f"{col_letter}{row}"].number_format = '0.00'

writer.close()
print("Peer Comparison Report generated successfully.")

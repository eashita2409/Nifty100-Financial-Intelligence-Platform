import pytest
import os
import pandas as pd
import numpy as np
import openpyxl
from pathlib import Path
from src.screener.engine import ScreenerEngine

@pytest.fixture
def db_path():
    project_root = Path(__file__).resolve().parent.parent.parent
    return project_root / "data" / "db" / "nifty100.db"

@pytest.fixture
def config_path():
    project_root = Path(__file__).resolve().parent.parent.parent
    return project_root / "config" / "screener_config.yaml"

@pytest.fixture
def test_df():
    # Mock data to test winsorization and sector normalization
    data = {
        'company_id': ['C1', 'C2', 'C3', 'C4', 'C5', 'C6', 'C7', 'C8', 'C9', 'C10'],
        'broad_sector': ['IT', 'IT', 'IT', 'IT', 'IT', 'Financials', 'Financials', 'Financials', 'Financials', 'Financials'],
        'roe': [5.0, 10.0, 15.0, 20.0, 45.0, 8.0, 12.0, 16.0, 20.0, 50.0],
        'roce': [6.0, 12.0, 18.0, 24.0, 50.0, 10.0, 14.0, 18.0, 22.0, 55.0],
        'cfo_quality_score': [0.8, 0.9, 1.0, 1.1, 1.5, 0.5, 0.7, 0.9, 1.0, 1.2],
        'revenue_cagr_5yr': [5.0, 8.0, 11.0, 14.0, 20.0, 6.0, 9.0, 12.0, 15.0, 25.0],
        'pat_cagr_5yr': [4.0, 7.0, 10.0, 13.0, 18.0, 5.0, 8.0, 11.0, 14.0, 22.0],
        'interest_coverage': [2.0, 4.0, 6.0, 8.0, 10.0, 1.5, 2.5, 3.5, 4.5, 5.5],
        'debt_equity': [0.1, 0.2, 0.3, 0.4, 0.5, 5.0, 6.0, 7.0, 8.0, 9.0],
        'debt_free_label': [0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
    }
    return pd.DataFrame(data)

def test_winsorization_and_normalization(test_df):
    engine = ScreenerEngine()
    df_scored = engine.calculate_composite_quality_score(test_df)
    
    # Verify normalization scales values between 0 and 100
    norm_cols = ['roe_norm', 'roce_norm', 'cfo_quality_norm', 'rev_cagr_norm', 'pat_cagr_norm']
    for col in norm_cols:
        vals = df_scored[col].dropna()
        assert vals.min() >= 0.0
        assert vals.max() <= 100.0

def test_financials_de_ignored_in_scoring(test_df):
    engine = ScreenerEngine()
    df_scored = engine.calculate_composite_quality_score(test_df)
    
    # Financials (C6-C10) should have NaN for debt_equity_norm
    financials = df_scored[df_scored['broad_sector'] == 'Financials']
    assert financials['debt_equity_norm'].isna().all()
    
    # IT sector (C1-C5) should have valid debt_equity_norm
    it_sector = df_scored[df_scored['broad_sector'] == 'IT']
    assert not it_sector['debt_equity_norm'].isna().any()

def test_composite_score_weight_contributions(test_df):
    engine = ScreenerEngine()
    df_scored = engine.calculate_composite_quality_score(test_df)
    
    # Let's verify score calculation manually for the first row (C1)
    # Weights: 35% Profitability, 30% Cash Quality, 20% Growth, 15% Leverage
    c1 = df_scored.iloc[0]
    profitability = np.mean([c1['roe_norm'], c1['roce_norm']])
    cash_quality = c1['cfo_quality_norm']
    growth = np.mean([c1['rev_cagr_norm'], c1['pat_cagr_norm']])
    leverage = np.mean([c1['icr_norm'], c1['debt_equity_norm']])
    
    expected_score = round(0.35 * profitability + 0.30 * cash_quality + 0.20 * growth + 0.15 * leverage, 2)
    assert c1['composite_quality_score'] == expected_score

def test_excel_export_file_generation(db_path, config_path, tmp_path):
    engine = ScreenerEngine(str(config_path))
    excel_file = tmp_path / "screener_output.xlsx"
    
    # Run export
    engine.export_to_excel(str(db_path), str(config_path), str(excel_file))
    
    # Verify file exists and is non-empty
    assert excel_file.exists()
    assert excel_file.stat().st_size > 0
    
    # Verify openpyxl can load it and it has the 6 sheets
    wb = openpyxl.load_workbook(excel_file)
    sheets = wb.sheetnames
    assert len(sheets) == 6
    assert "Quality Compounder" in sheets
    assert "Value Pick" in sheets
    assert "Growth Accelerator" in sheets
    assert "Dividend Champion" in sheets
    assert "Debt-Free Blue Chip" in sheets
    assert "Turnaround Watch" in sheets
    
    # Verify headers of the first sheet
    ws = wb["Quality Compounder"]
    headers = [cell.value for cell in ws[1]]
    assert headers[0] == "company_id"
    assert headers[1] == "company_name"
    assert headers[2] == "composite_quality_score"
    # Ensure there are 20 KPI columns (total of 23 columns)
    assert len(headers) == 23

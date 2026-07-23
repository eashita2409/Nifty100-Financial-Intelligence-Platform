import os
import yaml
import sqlite3
import pandas as pd
import numpy as np
from pathlib import Path
from typing import Optional, Dict, Any, List
import openpyxl
from openpyxl.styles import PatternFill, Font

class ScreenerEngine:
    def __init__(self, config_path: Optional[str] = None):
        self.config = {}
        self.presets = {}
        if config_path:
            self.load_config(config_path)

    def load_config(self, config_path: str):
        """Loads screener thresholds and presets from a YAML file."""
        path = Path(config_path)
        if not path.exists():
            raise FileNotFoundError(f"Configuration file not found: {path}")
        with open(path, 'r', encoding='utf-8') as f:
            data = yaml.safe_load(f)
        self.config = data.get('screener', {})
        self.presets = data.get('presets', {})

    def fetch_data(self, db_path: str, target_year: Optional[int] = None) -> pd.DataFrame:
        """Queries the SQLite database and builds the merged KPI dataframe for the target year (or latest if None)."""
        path = Path(db_path)
        if not path.exists():
            raise FileNotFoundError(f"Database not found: {path}")
            
        conn = sqlite3.connect(path)
        try:
            companies = pd.read_sql_query("SELECT id as company_id, company_name FROM companies", conn)
            sectors = pd.read_sql_query("SELECT company_id, broad_sector FROM sectors", conn)
            
            pnl_query = """
            SELECT 
                company_id, year, sales, net_profit, eps, 
                opm_percentage as operating_margin,
                sales as sales_val, net_profit as net_profit_val
            FROM profitandloss
            """
            pnl = pd.read_sql_query(pnl_query, conn)
            
            # Growth Metrics require historical data before filtering for latest year
            pnl = self._calc_growth(pnl, 'sales', 'revenue_growth')
            pnl = self._calc_growth(pnl, 'net_profit', 'pat_growth')
            pnl = self._calc_growth(pnl, 'eps', 'eps_growth')
            
            ratios_query = """
            SELECT 
                company_id, year, 
                return_on_equity_pct as roe, 
                return_on_capital_employed_pct as roce, 
                debt_to_equity as debt_equity, 
                free_cash_flow_cr as fcf, 
                revenue_cagr_5yr, 
                pat_cagr_5yr, 
                operating_profit_margin_pct as opm, 
                interest_coverage, 
                eps_cagr_5yr as eps_cagr, 
                asset_turnover, 
                debt_free_label,
                cfo_quality_score,
                net_profit_margin_pct as net_margin
            FROM financial_ratios
            """
            ratios = pd.read_sql_query(ratios_query, conn)
            
            bs_query = """
            SELECT company_id, year, other_asset, other_liabilities 
            FROM balancesheet
            """
            bs = pd.read_sql_query(bs_query, conn)
            
            mc_query = """
            SELECT 
                company_id, year, 
                market_cap_crore as market_cap, 
                pe_ratio as pe, 
                pb_ratio as pb, 
                dividend_yield_pct as dividend_yield,
                enterprise_value_crore as enterprise_value
            FROM market_cap
            """
            mc = pd.read_sql_query(mc_query, conn)
            
            # Growth for market cap
            mc = self._calc_growth(mc, 'market_cap', 'market_cap_growth')
            
            cashflow = pd.read_sql_query("SELECT company_id, year, net_cash_flow FROM cashflow", conn)
            cashflow = self._calc_growth(cashflow, 'net_cash_flow', 'cash_flow_growth')
            
            analysis = pd.read_sql_query("SELECT company_id, compounded_sales_growth as five_year_cagr FROM analysis WHERE compounded_sales_growth LIKE '5 Years%'", conn)
            
            # Merge temporal datasets
            if 'pnl' in locals() and hasattr(pnl, 'columns') and 'company_id' in pnl.columns: pnl['company_id'] = pnl['company_id'].astype(str).str.strip().str.upper()
            if 'ratios' in locals() and hasattr(ratios, 'columns') and 'company_id' in ratios.columns: ratios['company_id'] = ratios['company_id'].astype(str).str.strip().str.upper()
            merged = pnl.merge(ratios, on=['company_id', 'year'], how='outer')
            if 'merged' in locals() and hasattr(merged, 'columns') and 'company_id' in merged.columns: merged['company_id'] = merged['company_id'].astype(str).str.strip().str.upper()
            if 'bs' in locals() and hasattr(bs, 'columns') and 'company_id' in bs.columns: bs['company_id'] = bs['company_id'].astype(str).str.strip().str.upper()
            merged = merged.merge(bs, on=['company_id', 'year'], how='outer')
            if 'merged' in locals() and hasattr(merged, 'columns') and 'company_id' in merged.columns: merged['company_id'] = merged['company_id'].astype(str).str.strip().str.upper()
            if 'mc' in locals() and hasattr(mc, 'columns') and 'company_id' in mc.columns: mc['company_id'] = mc['company_id'].astype(str).str.strip().str.upper()
            merged = merged.merge(mc, on=['company_id', 'year'], how='outer')
            if 'merged' in locals() and hasattr(merged, 'columns') and 'company_id' in merged.columns: merged['company_id'] = merged['company_id'].astype(str).str.strip().str.upper()
            if 'cashflow' in locals() and hasattr(cashflow, 'columns') and 'company_id' in cashflow.columns: cashflow['company_id'] = cashflow['company_id'].astype(str).str.strip().str.upper()
            merged = merged.merge(cashflow, on=['company_id', 'year'], how='outer')
            
            # Filter by target_year or get latest valid year per company
            if target_year is not None:
                latest = merged[merged['year'] == target_year].copy()
            else:
                latest = merged.dropna(subset=['year']).sort_values('year').groupby('company_id').tail(1).copy()
            
            # Merge with metadata
            if 'latest' in locals() and hasattr(latest, 'columns') and 'company_id' in latest.columns: latest['company_id'] = latest['company_id'].astype(str).str.strip().str.upper()
            if 'companies' in locals() and hasattr(companies, 'columns') and 'company_id' in companies.columns: companies['company_id'] = companies['company_id'].astype(str).str.strip().str.upper()
            latest = latest.merge(companies, on='company_id', how='left')
            if 'latest' in locals() and hasattr(latest, 'columns') and 'company_id' in latest.columns: latest['company_id'] = latest['company_id'].astype(str).str.strip().str.upper()
            if 'sectors' in locals() and hasattr(sectors, 'columns') and 'company_id' in sectors.columns: sectors['company_id'] = sectors['company_id'].astype(str).str.strip().str.upper()
            latest = latest.merge(sectors, on='company_id', how='left')
            if 'latest' in locals() and hasattr(latest, 'columns') and 'company_id' in latest.columns: latest['company_id'] = latest['company_id'].astype(str).str.strip().str.upper()
            if 'analysis' in locals() and hasattr(analysis, 'columns') and 'company_id' in analysis.columns: analysis['company_id'] = analysis['company_id'].astype(str).str.strip().str.upper()
            latest = latest.merge(analysis, on='company_id', how='left')
            
            # Calculate current_ratio and quick_ratio
            latest['current_ratio'] = np.where(
                latest['other_liabilities'] == 0, 
                np.nan, 
                latest['other_asset'] / latest['other_liabilities']
            )
            latest['quick_ratio'] = latest['current_ratio']
            
            # Calculate PEG
            latest['peg'] = np.where(
                (latest['eps_growth'].isna()) | (latest['eps_growth'] <= 0) | (latest['pe'].isna()),
                np.nan,
                latest['pe'] / latest['eps_growth']
            )
            
            # Extract number for 5Y CAGR
            if 'five_year_cagr' in latest.columns:
                extracted = latest['five_year_cagr'].astype(str).str.extract(r':\s*(-?\d+\.?\d*)', expand=False)
                latest['five_year_cagr'] = pd.to_numeric(extracted, errors='coerce')
                
            return latest
        finally:
            conn.close()

    def _calc_growth(self, df: pd.DataFrame, col: str, new_col: str) -> pd.DataFrame:
        df = df.sort_values(['company_id', 'year'])
        df[f'{col}_prev'] = df.groupby('company_id')[col].shift(1)
        df[new_col] = np.where(
            df[f'{col}_prev'] == 0, 
            np.nan, 
            (df[col] - df[f'{col}_prev']) / df[f'{col}_prev'].abs() * 100
        )
        return df

    def clean_interest_coverage(self, val: Any, is_debt_free: bool) -> float:
        if is_debt_free:
            return float('inf')
        if val is None or pd.isna(val):
            return np.nan
        if isinstance(val, str):
            if val.strip().lower() == 'debt free':
                return float('inf')
            try:
                return float(val)
            except ValueError:
                return np.nan
        return float(val)

    def _winsorize_and_normalize(self, df: pd.DataFrame, col: str, ascending: bool) -> pd.Series:
        """Helper to winsorize (P10/P90) and normalize (0-100) a column sector-wise."""
        result = pd.Series(index=df.index, dtype=float)
        
        for sector, group in df.groupby('broad_sector', dropna=False):
            if col == 'debt_equity' and sector == 'Financials':
                result.loc[group.index] = np.nan
                continue
                
            vals = group[col].dropna()
            if len(vals) == 0:
                continue
                
            if len(vals) < 5:
                # Fall back to global percentile limits
                global_vals = df[col].dropna()
                p10 = global_vals.quantile(0.10) if len(global_vals) > 0 else 0.0
                p90 = global_vals.quantile(0.90) if len(global_vals) > 0 else 100.0
            else:
                p10 = vals.quantile(0.10)
                p90 = vals.quantile(0.90)
                
            if p90 == p10:
                p90 = p10 + 1e-5
                
            winsorized = np.clip(group[col], p10, p90)
            
            if ascending:
                norm = (winsorized - p10) / (p90 - p10) * 100
            else:
                norm = (p90 - winsorized) / (p90 - p10) * 100
                
            result.loc[group.index] = norm
            
        return result

    def calculate_composite_quality_score(self, df: pd.DataFrame) -> pd.DataFrame:
        """Calculates Composite Quality Score with winsorization (P10/P90) and sector-wise normalization."""
        df = df.copy()
        
        # Normalize interest coverage
        if 'interest_coverage' in df.columns:
            df['interest_coverage_clean'] = [
                self.clean_interest_coverage(row['interest_coverage'], row.get('debt_free_label') == 1)
                for _, row in df.iterrows()
            ]
        else:
            df['interest_coverage_clean'] = np.nan

        # Sector-wise winsorized normalization for individual sub-metrics
        df['roe_norm'] = self._winsorize_and_normalize(df, 'roe', True)
        df['roce_norm'] = self._winsorize_and_normalize(df, 'roce', True)
        df['cfo_quality_norm'] = self._winsorize_and_normalize(df, 'cfo_quality_score', True)
        df['rev_cagr_norm'] = self._winsorize_and_normalize(df, 'revenue_cagr_5yr', True)
        df['pat_cagr_norm'] = self._winsorize_and_normalize(df, 'pat_cagr_5yr', True)
        df['icr_norm'] = self._winsorize_and_normalize(df, 'interest_coverage_clean', True)
        df['debt_equity_norm'] = self._winsorize_and_normalize(df, 'debt_equity', False)

        # Dimension scores
        df['profitability_score'] = df[['roe_norm', 'roce_norm']].mean(axis=1)
        
        # Cash Quality (default to profitability if completely empty)
        df['cash_quality_score'] = df['cfo_quality_norm'].fillna(df['profitability_score'])
        
        # Growth
        df['growth_score'] = df[['rev_cagr_norm', 'pat_cagr_norm']].mean(axis=1)
        
        # Leverage: for financials, ignore debt_equity and use icr_norm alone
        df['leverage_score'] = np.where(
            df['broad_sector'] == 'Financials',
            df['icr_norm'],
            df[['icr_norm', 'debt_equity_norm']].mean(axis=1)
        )

        # Apply weights proportionally to non-NaN categories
        # Weights: 35% Profitability, 30% Cash Quality, 20% Growth, 15% Leverage
        cat_weights = {
            'profitability_score': 0.35,
            'cash_quality_score': 0.30,
            'growth_score': 0.20,
            'leverage_score': 0.15
        }
        
        total_weighted_scores = 0.0
        total_weights = 0.0
        for cat, weight in cat_weights.items():
            mask = df[cat].notna()
            total_weighted_scores = total_weighted_scores + df[cat].fillna(0) * weight
            total_weights = total_weights + np.where(mask, weight, 0)
            
        df['composite_quality_score'] = np.where(
            total_weights > 0,
            (total_weighted_scores / total_weights).round(2),
            0.0
        )
        
        return df

    def apply_filters(self, df: pd.DataFrame, config: Dict[str, Any]) -> pd.DataFrame:
        """Filters the dataframe according to the provided config rules."""
        df = df.copy()
        
        filter_mappings = {
            'roe_min': ('roe', 'gte'),
            'roce_min': ('roce', 'gte'),
            'debt_equity_max': ('debt_equity', 'lte_financial_bypass'),
            'fcf_min': ('fcf', 'gte'),
            'revenue_cagr_5y_min': ('revenue_cagr_5yr', 'gte'),
            'pat_cagr_5y_min': ('pat_cagr_5yr', 'gte'),
            'opm_min': ('operating_margin', 'gte'),
            'pe_max': ('pe', 'lte'),
            'pb_max': ('pb', 'lte'),
            'dividend_yield_min': ('dividend_yield', 'gte'),
            'interest_coverage_min': ('interest_coverage_clean', 'gte'),
            'market_cap_min': ('market_cap', 'gte'),
            'net_profit_min': ('net_profit_val', 'gte'),
            'eps_cagr_min': ('eps_cagr', 'gte'),
            'asset_turnover_min': ('asset_turnover', 'gte'),
            'sales_min': ('sales_val', 'gte')
        }
        
        # Ensure interest coverage is normalized before filtering
        if 'interest_coverage' in df.columns and 'interest_coverage_clean' not in df.columns:
            df['interest_coverage_clean'] = [
                self.clean_interest_coverage(row['interest_coverage'], row.get('debt_free_label') == 1)
                for _, row in df.iterrows()
            ]

        for key, value in config.items():
            if value is None:
                continue
            if key not in filter_mappings:
                continue
                
            col, op = filter_mappings[key]
            if col not in df.columns:
                continue
                
            if op == 'gte':
                df = df[df[col] >= value]
            elif op == 'lte':
                df = df[df[col] <= value]
            elif op == 'lte_financial_bypass':
                if 'broad_sector' in df.columns:
                    mask = (df['broad_sector'] == 'Financials') | (df[col] <= value)
                    df = df[mask]
                else:
                    df = df[df[col] <= value]
                    
        return df

    def run_screener(self, db_path: str, config_path: Optional[str] = None) -> pd.DataFrame:
        if config_path:
            self.load_config(config_path)
            
        df = self.fetch_data(db_path)
        df = self.calculate_composite_quality_score(df)
        df = self.apply_filters(df, self.config)
        
        if 'composite_quality_score' in df.columns:
            df = df.sort_values('composite_quality_score', ascending=False, na_position='last')
            
        return df

    def run_preset(self, preset_name: str, db_path: str, config_path: Optional[str] = None) -> pd.DataFrame:
        """Runs a named preset screener."""
        if config_path:
            self.load_config(config_path)
            
        if not hasattr(self, 'presets') or not self.presets or preset_name not in self.presets:
            raise ValueError(f"Preset '{preset_name}' not found. Available presets: {list(getattr(self, 'presets', {}).keys())}")
            
        preset_config = self.presets[preset_name]
        
        df = self.fetch_data(db_path)
        df = self.calculate_composite_quality_score(df)
        df = self.apply_filters(df, preset_config)
        
        if 'composite_quality_score' in df.columns:
            df = df.sort_values('composite_quality_score', ascending=False, na_position='last')
            
        return df

    def export_to_excel(self, db_path: str, config_path: str, excel_path: str):
        """Generates a formatted multi-sheet Excel file representing each preset."""
        self.load_config(config_path)
        
        wb = openpyxl.Workbook()
        # Remove default sheet
        default_sheet = wb.active
        wb.remove(default_sheet)
        
        green_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
        green_font = Font(color="155724", bold=True)
        red_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
        red_font = Font(color="721C24", bold=False)
        header_fill = PatternFill(start_color="E9ECEF", end_color="E9ECEF", fill_type="solid")
        header_font = Font(bold=True)
        
        # 20 target KPI columns
        kpi_columns = [
            'revenue_growth', 'pat_growth', 'eps_growth', 'roe', 'roce', 
            'debt_equity', 'current_ratio', 'quick_ratio', 'interest_coverage', 
            'operating_margin', 'net_margin', 'cash_flow_growth', 'market_cap_growth', 
            'peg', 'pe', 'pb', 'dividend_yield', 'enterprise_value', 'fcf', 
            'five_year_cagr'
        ]
        
        # Threshold key mapping
        threshold_map = {
            'roe': ('roe_min', 'gte'),
            'roce': ('roce_min', 'gte'),
            'debt_equity': ('debt_equity_max', 'lte_financial_bypass'),
            'fcf': ('fcf_min', 'gte'),
            'five_year_cagr': ('revenue_cagr_5y_min', 'gte'),
            'operating_margin': ('opm_min', 'gte'),
            'pe': ('pe_max', 'lte'),
            'pb': ('pb_max', 'lte'),
            'dividend_yield': ('dividend_yield_min', 'gte'),
            'interest_coverage': ('interest_coverage_min', 'gte')
        }

        for preset_name in self.presets.keys():
            res = self.run_preset(preset_name, db_path)
            
            # Create sheet
            # openpyxl limits sheet names to 31 chars
            sheet_title = preset_name[:31]
            ws = wb.create_sheet(title=sheet_title)
            
            # Write headers
            headers = ['company_id', 'company_name', 'composite_quality_score'] + kpi_columns
            ws.append(headers)
            
            # Format headers
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
            
            preset_config = self.presets[preset_name]
            
            # Write data rows
            for _, row in res.iterrows():
                row_data = [
                    row.get('company_id'),
                    row.get('company_name'),
                    row.get('composite_quality_score')
                ]
                for kpi in kpi_columns:
                    row_data.append(row.get(kpi))
                    
                ws.append(row_data)
                
                # Apply conditional formatting
                row_idx = ws.max_row
                is_financial = row.get('broad_sector') == 'Financials'
                
                for col_idx, kpi in enumerate(kpi_columns, start=4):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    val = cell.value
                    
                    if kpi not in threshold_map:
                        continue
                        
                    config_key, op = threshold_map[kpi]
                    threshold = preset_config.get(config_key)
                    
                    if threshold is None or val is None or pd.isna(val):
                        continue
                        
                    # Handle Interest Coverage special infinity value
                    if kpi == 'interest_coverage':
                        icr_val = self.clean_interest_coverage(val, row.get('debt_free_label') == 1)
                        if icr_val >= threshold:
                            cell.fill = green_fill
                            cell.font = green_font
                        else:
                            cell.fill = red_fill
                            cell.font = red_font
                        continue

                    # Handle filters
                    if op == 'gte':
                        if float(val) >= threshold:
                            cell.fill = green_fill
                            cell.font = green_font
                        else:
                            cell.fill = red_fill
                            cell.font = red_font
                    elif op == 'lte':
                        if float(val) <= threshold:
                            cell.fill = green_fill
                            cell.font = green_font
                        else:
                            cell.fill = red_fill
                            cell.font = red_font
                    elif op == 'lte_financial_bypass':
                        if is_financial:
                            cell.fill = green_fill
                            cell.font = green_font
                        elif float(val) <= threshold:
                            cell.fill = green_fill
                            cell.font = green_font
                        else:
                            cell.fill = red_fill
                            cell.font = red_font

            # Auto-fit columns
            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 3, 10)
                
        Path(excel_path).parent.mkdir(parents=True, exist_ok=True)
        wb.save(excel_path)
